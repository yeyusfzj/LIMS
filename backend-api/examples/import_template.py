"""
生成批量导入模板文件

创建 Excel 和 CSV 格式的导入模板，包含示例数据
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import os


def create_excel_template():
    """创建 Excel 导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "检测结果导入"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = [
        "sampleId",
        "testItemId",
        "parameter",
        "value",
        "textValue",
        "unit",
        "method",
        "instrumentId"
    ]
    
    # 字段说明
    descriptions = [
        "样品ID（必填）",
        "检测项ID（必填）",
        "检测参数（必填）",
        "数值结果（可选）",
        "文本结果（可选）",
        "单位（可选）",
        "检测方法（必填）",
        "仪器ID（可选）"
    ]
    
    # 写入表头
    for col, (header, desc) in enumerate(zip(headers, descriptions), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
        # 添加注释（使用 Comment 对象）
        from openpyxl.comments import Comment
        cell.comment = Comment(desc, "System")
    
    # 写入说明行
    ws.cell(row=2, column=1, value="说明：")
    ws.cell(row=2, column=2, value="value 和 textValue 至少填写一个")
    
    # 写入示例数据
    examples = [
        ["sample-001", "item-001", "pH", 7.5, "", "pH", "GB/T 5750.4-2006", "instrument-001"],
        ["sample-001", "item-002", "外观", "", "无色透明", "", "目测法", ""],
        ["sample-001", "item-003", "浊度", 2.3, "", "NTU", "GB/T 5750.4-2006", "instrument-002"],
        ["sample-002", "item-004", "色度", 5, "", "度", "GB/T 5750.4-2006", "instrument-003"],
        ["sample-002", "item-005", "气味", "", "无异味", "", "嗅觉法", ""],
    ]
    
    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    column_widths = [15, 15, 15, 12, 15, 10, 25, 18]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 保存文件
    output_dir = os.path.dirname(__file__)
    excel_path = os.path.join(output_dir, "results_import_template.xlsx")
    wb.save(excel_path)
    
    print(f"✓ Excel 模板已创建: {excel_path}")
    return excel_path


def create_csv_template():
    """创建 CSV 导入模板"""
    output_dir = os.path.dirname(__file__)
    csv_path = os.path.join(output_dir, "results_import_template.csv")
    
    # 表头
    headers = [
        "sampleId",
        "testItemId",
        "parameter",
        "value",
        "textValue",
        "unit",
        "method",
        "instrumentId"
    ]
    
    # 示例数据
    examples = [
        ["sample-001", "item-001", "pH", "7.5", "", "pH", "GB/T 5750.4-2006", "instrument-001"],
        ["sample-001", "item-002", "外观", "", "无色透明", "", "目测法", ""],
        ["sample-001", "item-003", "浊度", "2.3", "", "NTU", "GB/T 5750.4-2006", "instrument-002"],
        ["sample-002", "item-004", "色度", "5", "", "度", "GB/T 5750.4-2006", "instrument-003"],
        ["sample-002", "item-005", "气味", "", "无异味", "", "嗅觉法", ""],
    ]
    
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(examples)
    
    print(f"✓ CSV 模板已创建: {csv_path}")
    return csv_path


def create_readme():
    """创建使用说明文档"""
    output_dir = os.path.dirname(__file__)
    readme_path = os.path.join(output_dir, "IMPORT_TEMPLATE_README.md")
    
    content = """# 批量导入模板使用说明

## 文件格式

支持以下两种文件格式：
- Excel (.xlsx, .xls)
- CSV (.csv)

## 字段说明

| 字段名 | 类型 | 必填 | 说明 |
|--------|------|------|------|
| sampleId | 字符串 | 是 | 样品 ID，必须在系统中存在 |
| testItemId | 字符串 | 是 | 检测项 ID |
| parameter | 字符串 | 是 | 检测参数名称 |
| value | 数值 | 否* | 数值型结果 |
| textValue | 字符串 | 否* | 文本型结果 |
| unit | 字符串 | 否 | 单位 |
| method | 字符串 | 是 | 检测方法 |
| instrumentId | 字符串 | 否 | 仪器 ID |

**注意**: value 和 textValue 至少需要填写一个

## 使用步骤

1. **下载模板**
   - Excel 模板: `results_import_template.xlsx`
   - CSV 模板: `results_import_template.csv`

2. **填写数据**
   - 保留表头行（第一行）
   - 从第二行开始填写数据
   - 确保必填字段不为空
   - 数值型结果填写在 value 列
   - 文本型结果填写在 textValue 列

3. **上传导入**
   - 访问系统的批量导入页面
   - 选择填写好的文件
   - 点击上传按钮
   - 等待导入完成

4. **查看结果**
   - 系统会显示导入统计信息
   - 成功数量、失败数量
   - 如有错误，会显示详细的错误信息（行号、字段、原因）

## 示例数据

### 数值型结果示例

```
sampleId,testItemId,parameter,value,textValue,unit,method,instrumentId
sample-001,item-001,pH,7.5,,pH,GB/T 5750.4-2006,instrument-001
sample-001,item-003,浊度,2.3,,NTU,GB/T 5750.4-2006,instrument-002
```

### 文本型结果示例

```
sampleId,testItemId,parameter,value,textValue,unit,method,instrumentId
sample-001,item-002,外观,,无色透明,,目测法,
sample-002,item-005,气味,,无异味,,嗅觉法,
```

## 常见错误

### 1. 样品不存在
**错误信息**: "样品不存在"
**解决方法**: 确保 sampleId 在系统中已经注册

### 2. 必填字段为空
**错误信息**: "XXX 不能为空"
**解决方法**: 填写必填字段（sampleId, testItemId, parameter, method）

### 3. 数值格式不正确
**错误信息**: "数值格式不正确"
**解决方法**: 确保 value 列填写的是有效的数字

### 4. 缺少结果值
**错误信息**: "数值结果或文本结果至少需要提供一个"
**解决方法**: 至少填写 value 或 textValue 中的一个

## 性能建议

- 单次导入建议不超过 1000 条记录
- 文件大小不超过 10MB
- 对于大批量数据，建议分批导入

## API 调用示例

### Python

```python
import httpx

async with httpx.AsyncClient() as client:
    with open('results.xlsx', 'rb') as f:
        files = {'file': ('results.xlsx', f)}
        response = await client.post(
            'http://localhost:8000/api/v1/results/import',
            files=files,
            headers={'Authorization': f'Bearer {token}'}
        )
        print(response.json())
```

### cURL

```bash
curl -X POST "http://localhost:8000/api/v1/results/import" \\
  -H "Authorization: Bearer YOUR_TOKEN" \\
  -F "file=@results.xlsx"
```

## 技术支持

如有问题，请联系系统管理员或查看系统文档。
"""
    
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f"✓ 使用说明已创建: {readme_path}")
    return readme_path


def main():
    """主函数"""
    print("=" * 60)
    print("生成批量导入模板文件")
    print("=" * 60)
    print()
    
    try:
        # 创建模板文件
        excel_path = create_excel_template()
        csv_path = create_csv_template()
        readme_path = create_readme()
        
        print()
        print("=" * 60)
        print("✓ 所有模板文件已生成！")
        print("=" * 60)
        print()
        print("生成的文件：")
        print(f"  1. {excel_path}")
        print(f"  2. {csv_path}")
        print(f"  3. {readme_path}")
        
    except Exception as e:
        print(f"✗ 生成失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
