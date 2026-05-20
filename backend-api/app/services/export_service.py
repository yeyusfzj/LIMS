"""
数据导出服务

支持将数据导出为 Excel 和 CSV 格式，支持异步导出任务管理
"""
from typing import List, Dict, Any, Optional, Literal
from datetime import datetime, timedelta
from enum import Enum
import os
import csv
import uuid
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from app.core.logging import logger


class ExportFormat(str, Enum):
    """导出格式枚举"""
    EXCEL = "excel"
    CSV = "csv"


class ExportStatus(str, Enum):
    """导出任务状态枚举"""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"


class ExportTask:
    """导出任务"""
    def __init__(
        self,
        task_id: str,
        format: ExportFormat,
        status: ExportStatus = ExportStatus.PENDING,
        file_path: Optional[str] = None,
        download_url: Optional[str] = None,
        error: Optional[str] = None,
        created_at: Optional[datetime] = None,
        completed_at: Optional[datetime] = None,
        expires_at: Optional[datetime] = None
    ):
        self.task_id = task_id
        self.format = format
        self.status = status
        self.file_path = file_path
        self.download_url = download_url
        self.error = error
        self.created_at = created_at or datetime.utcnow()
        self.completed_at = completed_at
        self.expires_at = expires_at
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "taskId": self.task_id,
            "format": self.format,
            "status": self.status,
            "filePath": self.file_path,
            "downloadUrl": self.download_url,
            "error": self.error,
            "createdAt": self.created_at.isoformat() if self.created_at else None,
            "completedAt": self.completed_at.isoformat() if self.completed_at else None,
            "expiresAt": self.expires_at.isoformat() if self.expires_at else None
        }


class ExportService:
    """导出服务类"""
    
    # 导出目录
    EXPORT_DIR = Path("exports")
    
    # 文件过期时间（小时）
    FILE_EXPIRY_HOURS = 24
    
    # 导出任务存储（生产环境应使用 Redis）
    _export_tasks: Dict[str, ExportTask] = {}
    
    @classmethod
    def initialize(cls) -> None:
        """初始化导出目录"""
        try:
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            logger.info(f"Export directory initialized: {cls.EXPORT_DIR}")
        except Exception as e:
            logger.error(f"Failed to initialize export directory: {str(e)}")
            raise
    
    @classmethod
    async def create_export_task(
        cls,
        format: ExportFormat,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        filename: Optional[str] = None
    ) -> ExportTask:
        """
        创建导出任务
        
        Args:
            format: 导出格式
            data: 要导出的数据
            columns: 列名列表（可选，如果不提供则从数据中提取）
            filename: 文件名（可选）
            
        Returns:
            ExportTask: 导出任务对象
        """
        try:
            # 生成任务 ID
            task_id = str(uuid.uuid4())
            
            # 创建任务
            task = ExportTask(
                task_id=task_id,
                format=format,
                status=ExportStatus.PENDING
            )
            
            # 存储任务
            cls._export_tasks[task_id] = task
            
            # 异步处理导出（这里简化为同步处理）
            await cls._process_export_task(task, data, columns, filename)
            
            return task
            
        except Exception as e:
            logger.error(f"Failed to create export task: {str(e)}")
            raise
    
    @classmethod
    async def _process_export_task(
        cls,
        task: ExportTask,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        filename: Optional[str] = None
    ) -> None:
        """
        处理导出任务
        
        Args:
            task: 导出任务
            data: 要导出的数据
            columns: 列名列表
            filename: 文件名
        """
        try:
            task.status = ExportStatus.PROCESSING
            logger.info(f"Processing export task: {task.task_id}, format={task.format}")
            
            # 根据格式导出
            if task.format == ExportFormat.EXCEL:
                file_path = await cls.export_to_excel(data, columns, filename)
            elif task.format == ExportFormat.CSV:
                file_path = await cls.export_to_csv(data, columns, filename)
            else:
                raise ValueError(f"Unsupported export format: {task.format}")
            
            # 生成下载链接
            file_name = Path(file_path).name
            download_url = f"/api/v1/export/download/{file_name}"
            expires_at = datetime.utcnow() + timedelta(hours=cls.FILE_EXPIRY_HOURS)
            
            # 更新任务状态
            task.status = ExportStatus.COMPLETED
            task.file_path = file_path
            task.download_url = download_url
            task.expires_at = expires_at
            task.completed_at = datetime.utcnow()
            
            logger.info(f"Export task completed: {task.task_id}, file={file_path}")
            
        except Exception as e:
            task.status = ExportStatus.FAILED
            task.error = str(e)
            logger.error(f"Export task failed: {task.task_id}, error={str(e)}")
    
    @classmethod
    async def get_export_task(cls, task_id: str) -> Optional[ExportTask]:
        """
        获取导出任务
        
        Args:
            task_id: 任务 ID
            
        Returns:
            Optional[ExportTask]: 导出任务对象，如果不存在则返回 None
        """
        return cls._export_tasks.get(task_id)
    
    @classmethod
    async def export_to_excel(
        cls,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        导出数据为 Excel 格式
        
        Args:
            data: 要导出的数据列表
            columns: 列名列表（可选，如果不提供则从数据中提取）
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            # 确保导出目录存在
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"export_{timestamp}.xlsx"
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
            file_path = cls.EXPORT_DIR / file_name
            
            # 创建工作簿
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "数据"
            
            if not data:
                # 空数据
                workbook.save(file_path)
                logger.info(f"Excel export completed (empty): {file_path}")
                return str(file_path)
            
            # 确定列名
            if columns is None:
                columns = list(data[0].keys())
            
            # 写入表头
            for col_num, header in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据
            for row_num, item in enumerate(data, 2):
                for col_num, column in enumerate(columns, 1):
                    value = item.get(column)
                    
                    # 处理日期时间
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    
                    worksheet.cell(row=row_num, column=col_num).value = value
            
            # 调整列宽
            for col_num in range(1, len(columns) + 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = 15
            
            # 保存文件
            workbook.save(file_path)
            
            logger.info(f"Excel export completed: {file_path}, rows={len(data)}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {str(e)}")
            raise
    
    @classmethod
    async def export_to_csv(
        cls,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        导出数据为 CSV 格式
        
        Args:
            data: 要导出的数据列表
            columns: 列名列表（可选，如果不提供则从数据中提取）
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            # 确保导出目录存在
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"export_{timestamp}.csv"
            if not file_name.endswith('.csv'):
                file_name += '.csv'
            file_path = cls.EXPORT_DIR / file_name
            
            if not data:
                # 空数据
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    pass
                logger.info(f"CSV export completed (empty): {file_path}")
                return str(file_path)
            
            # 确定列名
            if columns is None:
                columns = list(data[0].keys())
            
            # 写入 CSV 文件
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                
                # 写入表头
                writer.writeheader()
                
                # 写入数据
                for item in data:
                    row = {}
                    for column in columns:
                        value = item.get(column)
                        
                        # 处理日期时间
                        if isinstance(value, datetime):
                            value = value.strftime("%Y-%m-%d %H:%M:%S")
                        
                        row[column] = value if value is not None else ""
                    
                    writer.writerow(row)
            
            logger.info(f"CSV export completed: {file_path}, rows={len(data)}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export to CSV: {str(e)}")
            raise
    
    @classmethod
    async def get_export_file(cls, filename: str) -> Optional[str]:
        """
        获取导出文件路径
        
        Args:
            filename: 文件名
            
        Returns:
            Optional[str]: 文件路径，如果文件不存在或不安全则返回 None
        """
        try:
            file_path = cls.EXPORT_DIR / filename
            
            # 检查文件是否存在
            if not file_path.exists():
                return None
            
            # 安全检查：确保文件在导出目录内
            resolved_path = file_path.resolve()
            resolved_export_dir = cls.EXPORT_DIR.resolve()
            
            if not str(resolved_path).startswith(str(resolved_export_dir)):
                logger.warning(f"Attempted to access file outside export directory: {filename}")
                return None
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to get export file: {str(e)}")
            return None
    
    @classmethod
    async def cleanup_expired_files(cls) -> int:
        """
        清理过期文件
        
        Returns:
            int: 删除的文件数量
        """
        try:
            if not cls.EXPORT_DIR.exists():
                return 0
            
            deleted_count = 0
            now = datetime.utcnow()
            max_age = timedelta(hours=cls.FILE_EXPIRY_HOURS)
            
            for file_path in cls.EXPORT_DIR.iterdir():
                if file_path.is_file():
                    # 获取文件修改时间
                    mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    file_age = now - mtime
                    
                    if file_age > max_age:
                        file_path.unlink()
                        deleted_count += 1
                        logger.debug(f"Deleted expired export file: {file_path.name}")
            
            if deleted_count > 0:
                logger.info(f"Cleaned up {deleted_count} expired export files")
            
            return deleted_count
            
        except Exception as e:
            logger.error(f"Failed to cleanup expired files: {str(e)}")
            return 0
    
    @classmethod
    async def export_audit_tasks_to_excel(
        cls,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        导出审核任务列表为 Excel 格式
        
        Args:
            data: 审核任务数据列表
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            # 确保导出目录存在
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"audit_tasks_{timestamp}.xlsx"
            file_path = cls.EXPORT_DIR / file_name
            
            # 创建工作簿
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "审核任务"
            
            # 定义表头
            headers = [
                "任务ID",
                "样品ID",
                "样品编号",
                "样品名称",
                "样品类型",
                "审核级别",
                "审核人员ID",
                "状态",
                "决策",
                "审核意见",
                "提交时间",
                "完成时间"
            ]
            
            # 写入表头
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据
            for row_num, task in enumerate(data, 2):
                sample = task.get("sample", {}) or {}
                
                worksheet.cell(row=row_num, column=1).value = task.get("id", "")
                worksheet.cell(row=row_num, column=2).value = task.get("sampleId", "")
                worksheet.cell(row=row_num, column=3).value = sample.get("sampleNumber", "")
                worksheet.cell(row=row_num, column=4).value = sample.get("sampleName", "")
                worksheet.cell(row=row_num, column=5).value = sample.get("sampleType", "")
                worksheet.cell(row=row_num, column=6).value = task.get("level", "")
                worksheet.cell(row=row_num, column=7).value = task.get("auditorId", "")
                worksheet.cell(row=row_num, column=8).value = task.get("status", "")
                worksheet.cell(row=row_num, column=9).value = task.get("decision", "") if task.get("decision") else ""
                worksheet.cell(row=row_num, column=10).value = task.get("comments", "") if task.get("comments") else ""
                
                # 格式化时间
                submitted_at = task.get("submittedAt")
                if submitted_at:
                    if isinstance(submitted_at, str):
                        worksheet.cell(row=row_num, column=11).value = submitted_at
                    else:
                        worksheet.cell(row=row_num, column=11).value = submitted_at.strftime("%Y-%m-%d %H:%M:%S")
                
                completed_at = task.get("completedAt")
                if completed_at:
                    if isinstance(completed_at, str):
                        worksheet.cell(row=row_num, column=12).value = completed_at
                    else:
                        worksheet.cell(row=row_num, column=12).value = completed_at.strftime("%Y-%m-%d %H:%M:%S")
            
            # 调整列宽
            column_widths = [15, 15, 15, 20, 15, 12, 15, 12, 12, 30, 20, 20]
            for col_num, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = width
            
            # 保存文件
            workbook.save(file_path)
            
            logger.info(f"Audit tasks exported to Excel: {file_path}, rows={len(data)}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export audit tasks to Excel: {str(e)}")
            raise
    
    @classmethod
    async def export_workload_to_excel(
        cls,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        导出工作量统计为 Excel 格式
        
        Args:
            data: 工作量统计数据
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"workload_statistics_{timestamp}.xlsx"
            file_path = cls.EXPORT_DIR / file_name
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "工作量统计"
            
            # 表头
            headers = ["审核人员", "总任务数", "已完成", "待处理"]
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 数据
            for row_num, item in enumerate(data, 2):
                worksheet.cell(row=row_num, column=1).value = item.get("auditorName", "")
                worksheet.cell(row=row_num, column=2).value = item.get("totalTasks", 0)
                worksheet.cell(row=row_num, column=3).value = item.get("completedTasks", 0)
                worksheet.cell(row=row_num, column=4).value = item.get("pendingTasks", 0)
            
            # 列宽
            for col_num, width in enumerate([15, 12, 12, 12], 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = width
            
            workbook.save(file_path)
            
            logger.info(f"Workload statistics exported: {file_path}, rows={len(data)}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export workload statistics: {str(e)}")
            raise
    
    @classmethod
    async def export_pass_rate_to_excel(
        cls,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        导出通过率统计为 Excel 格式
        
        Args:
            data: 通过率统计数据
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"pass_rate_statistics_{timestamp}.xlsx"
            file_path = cls.EXPORT_DIR / file_name
            
            workbook = Workbook()
            
            # 整体统计工作表
            overall_sheet = workbook.active
            overall_sheet.title = "整体统计"
            
            overall_data = [
                ["指标", "数值"],
                ["总任务数", data.get("overall", {}).get("total", 0)],
                ["通过数", data.get("overall", {}).get("passed", 0)],
                ["退回数", data.get("overall", {}).get("rejected", 0)],
                ["通过率(%)", f"{data.get('overall', {}).get('passRate', 0):.2f}"]
            ]
            
            for row_num, row_data in enumerate(overall_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = overall_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            overall_sheet.column_dimensions['A'].width = 15
            overall_sheet.column_dimensions['B'].width = 15
            
            # 按级别统计工作表
            level_sheet = workbook.create_sheet("按级别统计")
            level_data = [["审核级别", "总任务数", "通过数", "通过率(%)"]]
            
            for item in data.get("byLevel", []):
                level_data.append([
                    f"第{item.get('level', '')}级",
                    item.get("total", 0),
                    item.get("passed", 0),
                    f"{item.get('passRate', 0):.2f}"
                ])
            
            for row_num, row_data in enumerate(level_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = level_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for col_num, width in enumerate([12, 12, 12, 15], 1):
                level_sheet.column_dimensions[get_column_letter(col_num)].width = width
            
            # 按样品类型统计工作表
            type_sheet = workbook.create_sheet("按样品类型统计")
            type_data = [["样品类型", "总任务数", "通过数", "通过率(%)"]]
            
            for item in data.get("bySampleType", []):
                type_data.append([
                    item.get("sampleType", ""),
                    item.get("total", 0),
                    item.get("passed", 0),
                    f"{item.get('passRate', 0):.2f}"
                ])
            
            for row_num, row_data in enumerate(type_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = type_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for col_num, width in enumerate([15, 12, 12, 15], 1):
                type_sheet.column_dimensions[get_column_letter(col_num)].width = width
            
            workbook.save(file_path)
            
            logger.info(f"Pass rate statistics exported: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export pass rate statistics: {str(e)}")
            raise
    
    @classmethod
    async def export_duration_to_excel(
        cls,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        导出时效性统计为 Excel 格式
        
        Args:
            data: 时效性统计数据
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"duration_statistics_{timestamp}.xlsx"
            file_path = cls.EXPORT_DIR / file_name
            
            workbook = Workbook()
            
            # 统计指标工作表
            metrics_sheet = workbook.active
            metrics_sheet.title = "统计指标"
            
            overall = data.get("overall", {})
            metrics_data = [
                ["指标", "数值(小时)"],
                ["平均时长", f"{overall.get('averageDuration', 0):.2f}"],
                ["中位数", f"{overall.get('medianDuration', 0):.2f}"],
                ["最短时长", f"{overall.get('minDuration', 0):.2f}"],
                ["最长时长", f"{overall.get('maxDuration', 0):.2f}"],
                ["超时任务数", overall.get("overtimeTasks", 0)],
                ["超时率(%)", f"{overall.get('overtimeRate', 0):.2f}"]
            ]
            
            for row_num, row_data in enumerate(metrics_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = metrics_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            metrics_sheet.column_dimensions['A'].width = 15
            metrics_sheet.column_dimensions['B'].width = 15
            
            # 时长分布工作表
            distribution_sheet = workbook.create_sheet("时长分布")
            distribution_data = [["时长范围", "任务数量"]]
            
            for item in data.get("distribution", []):
                distribution_data.append([
                    item.get("range", ""),
                    item.get("count", 0)
                ])
            
            for row_num, row_data in enumerate(distribution_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = distribution_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            distribution_sheet.column_dimensions['A'].width = 20
            distribution_sheet.column_dimensions['B'].width = 12
            
            workbook.save(file_path)
            
            logger.info(f"Duration statistics exported: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export duration statistics: {str(e)}")
            raise
    
    @classmethod
    async def export_issues_to_excel(
        cls,
        data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        导出问题分类统计为 Excel 格式
        
        Args:
            data: 问题分类统计数据
            filename: 文件名（可选）
            
        Returns:
            str: 导出文件的路径
        """
        try:
            cls.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%f")[:-3] + "Z"
            file_name = filename or f"issues_statistics_{timestamp}.xlsx"
            file_path = cls.EXPORT_DIR / file_name
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "问题分类统计"
            
            # 表头
            headers = ["排名", "退回原因", "出现次数", "占比(%)"]
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 数据
            by_reason = data.get("byReason", [])
            total = 0
            
            for row_num, item in enumerate(by_reason, 2):
                worksheet.cell(row=row_num, column=1).value = row_num - 1
                worksheet.cell(row=row_num, column=2).value = item.get("reason", "")
                count = item.get("count", 0)
                worksheet.cell(row=row_num, column=3).value = count
                worksheet.cell(row=row_num, column=4).value = f"{item.get('percentage', 0):.2f}"
                total += count
            
            # 总计行
            if by_reason:
                total_row = len(by_reason) + 3
                worksheet.cell(row=total_row, column=1).value = "总计"
                worksheet.cell(row=total_row, column=1).font = Font(bold=True)
                worksheet.cell(row=total_row, column=3).value = total
                worksheet.cell(row=total_row, column=3).font = Font(bold=True)
                worksheet.cell(row=total_row, column=4).value = "100.00"
                worksheet.cell(row=total_row, column=4).font = Font(bold=True)
            
            # 列宽
            for col_num, width in enumerate([8, 30, 12, 12], 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = width
            
            workbook.save(file_path)
            
            logger.info(f"Issues statistics exported: {file_path}, issues={len(by_reason)}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to export issues statistics: {str(e)}")
            raise


# 创建服务实例
export_service = ExportService()

