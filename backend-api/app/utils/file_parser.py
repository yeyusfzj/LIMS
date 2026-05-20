"""
文件解析工具

支持 CSV、Excel 格式的文件解析
验证需求：3.2, 11.8
"""
import csv
import io
from typing import List, Dict, Any, Optional
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class FileParser:
    """文件解析器类"""

    def parse_csv(self, content: bytes) -> List[Dict[str, Any]]:
        """
        解析 CSV 文件
        
        Args:
            content: 文件内容（字节）
            
        Returns:
            解析后的数据行列表
            
        Raises:
            ValueError: 解析失败
        """
        try:
            # 尝试 UTF-8 编码
            try:
                text = content.decode('utf-8')
            except UnicodeDecodeError:
                # 尝试 GBK 编码（中文 Excel 导出的 CSV）
                text = content.decode('gbk')
            
            # 解析 CSV
            reader = csv.DictReader(io.StringIO(text))
            records = list(reader)
            
            logger.info(f"CSV parsed successfully, {len(records)} records")
            return records
            
        except Exception as e:
            logger.error(f"Failed to parse CSV: {str(e)}")
            raise ValueError(f"CSV 解析失败: {str(e)}")

    def parse_excel(self, content: bytes) -> List[Dict[str, Any]]:
        """
        解析 Excel 文件
        
        Args:
            content: 文件内容（字节）
            
        Returns:
            解析后的数据行列表
            
        Raises:
            ValueError: 解析失败
        """
        try:
            # 加载工作簿
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            
            # 获取第一个工作表
            sheet = workbook.active
            
            # 读取表头（第一行）
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            # 读取数据行
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # 转换为字符串，处理 None 值
                        record[headers[i]] = str(value).strip() if value is not None else ''
                
                # 跳过空行
                if any(record.values()):
                    records.append(record)
            
            workbook.close()
            
            logger.info(f"Excel parsed successfully, {len(records)} records")
            return records
            
        except Exception as e:
            logger.error(f"Failed to parse Excel: {str(e)}")
            raise ValueError(f"Excel 解析失败: {str(e)}")

    def parse_file(self, content: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        根据文件扩展名解析文件
        
        Args:
            content: 文件内容（字节）
            filename: 文件名
            
        Returns:
            解析后的数据行列表
            
        Raises:
            ValueError: 不支持的文件格式或解析失败
        """
        ext = filename.lower().split('.')[-1]
        
        if ext == 'csv':
            return self.parse_csv(content)
        elif ext in ['xlsx', 'xls']:
            return self.parse_excel(content)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")


# 创建全局实例
file_parser = FileParser()
